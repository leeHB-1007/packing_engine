from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from master_loader import (
    load_master_workbook,
    prepare_boxes_for_engine,
    prepare_products_for_engine,
)


DEFAULT_MASTER = Path("data/packing_engine_normalized_masters_ko_json_schema_fixed.xlsx")

HEADER_TITLES = [
    "표기",
    "박스번호",
    "품명",
    "포장단위",
    "박스사이즈(Cm)",
    "박스수량",
    "낱개수량",
    "중량(Kg)",
    "합계(kg)",
    "비고",
]


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(value)
    except Exception:
        return default


def _to_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(round(float(value)))
    except Exception:
        return default


def _fmt_num(value: Any) -> str:
    num = _to_float(value, 0.0)
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.1f}".rstrip("0").rstrip(".")


def _fmt_weight(value: Any) -> float | None:
    if value in (None, ""):
        return None
    num = _to_float(value, 0.0)
    return round(num, 3)


def _normalize_calc_unit(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"package", "pkg", "pack"}:
        return "package"
    return "item"


def _format_box_size(outer_dims: Any, trimmed_outer_height: Any = None) -> str:
    if not isinstance(outer_dims, (list, tuple)) or len(outer_dims) < 3:
        return ""

    length = _to_float(outer_dims[0], 0.0)
    width = _to_float(outer_dims[1], 0.0)
    height = _to_float(outer_dims[2], 0.0)

    trimmed_height = _to_float(trimmed_outer_height, 0.0)
    if trimmed_height > 0 and trimmed_height < height:
        height = trimmed_height

    return f"{_fmt_num(length)} X {_fmt_num(width)} X {_fmt_num(height)}"


def _build_master_context(master_path: str | Path | None = None) -> Dict[str, dict]:
    master_file = Path(master_path) if master_path else DEFAULT_MASTER
    load_result = load_master_workbook(master_file)

    prepared_products = prepare_products_for_engine(
        load_result["products"],
        load_result["fullboxes"],
    )
    prepared_boxes = prepare_boxes_for_engine(
        load_result["boxes"],
    )

    product_lookup: Dict[str, dict] = {}
    for _, row in prepared_products.iterrows():
        product_name = str(row.get("국문상품명", "") or "").strip()
        if not product_name:
            continue

        product_lookup[product_name] = {
            "product_code": str(row.get("상품코드", "") or "").strip(),
            "unit_weight_kg": _to_float(row.get("개당중량(kg)"), 0.0),
        }

    box_lookup: Dict[str, dict] = {}
    for _, row in prepared_boxes.iterrows():
        box_code = str(row.get("박스코드", "") or "").strip()
        if not box_code:
            continue

        box_lookup[box_code] = {
            "box_name": str(row.get("박스명", "") or "").strip(),
            "outer_size_cm": (
                _to_float(row.get("외경가로(cm)"), 0.0),
                _to_float(row.get("외경세로(cm)"), 0.0),
                _to_float(row.get("외경높이(cm)"), 0.0),
            ),
            "box_weight_kg": _to_float(row.get("박스중량(kg)"), 0.0),
            "weight_mode": "box_only",
        }

    fullboxes_df = load_result["fullboxes"]
    for _, row in fullboxes_df.iterrows():
        box_code = str(row.get("완박스박스코드", "") or "").strip()
        if not box_code:
            continue

        box_lookup[box_code] = {
            "box_name": str(row.get("완박스박스명", "") or "").strip(),
            "outer_size_cm": (
                _to_float(row.get("완박스가로(cm)"), 0.0),
                _to_float(row.get("완박스세로(cm)"), 0.0),
                _to_float(row.get("완박스높이(cm)"), 0.0),
            ),
            # fullboxes_master weight is treated as gross weight for one full box.
            "box_weight_kg": _to_float(row.get("완박스중량(kg)"), 0.0),
            "weight_mode": "fullbox_gross",
        }

    return {
        "product_lookup": product_lookup,
        "box_lookup": box_lookup,
    }


def _build_code_map(
    engine_result: Dict[str, Any],
    product_lookup: Dict[str, dict],
) -> Dict[str, str]:
    code_map = {
        product_name: payload.get("product_code", "")
        for product_name, payload in product_lookup.items()
        if payload.get("product_code")
    }

    for row in engine_result.get("match_result", []) or []:
        if isinstance(row, dict):
            matched_name = row.get("matched_name") or row.get("product_name")
            product_code = row.get("product_code") or row.get("matched_code")
        else:
            matched_name = getattr(row, "matched_name", "") or getattr(row, "product_name", "")
            product_code = getattr(row, "product_code", "") or getattr(row, "matched_code", "")

        matched_name = str(matched_name or "").strip()
        product_code = str(product_code or "").strip()
        if matched_name and product_code:
            code_map[matched_name] = product_code

    return code_map


def _new_item_row(
    *,
    product_code: str,
    product_name: str,
    packing_unit: Any,
) -> Dict[str, Any]:
    return {
        "product_code": str(product_code or "").strip(),
        "product_name": str(product_name or "").strip(),
        "packing_unit": packing_unit,
    }


def _new_export_group(
    *,
    box_no: int,
    box_size_cm: str,
    box_count: int,
    each_qty: int,
    weight_kg: Any,
    total_weight_kg: Any,
    item_rows: List[Dict[str, Any]],
    note: str = "",
) -> Dict[str, Any]:
    return {
        "box_size_cm": box_size_cm,
        "box_no_start": int(box_no),
        "box_no_end": int(box_no),
        "box_count": int(box_count),
        "each_qty": int(each_qty),
        "weight_kg": _fmt_weight(weight_kg),
        "total_weight_kg": _fmt_weight(total_weight_kg),
        "item_rows": item_rows,
        "note": str(note or "").strip(),
    }


def _build_fullbox_groups(
    fullbox_result: Dict[str, Any],
    box_lookup: Dict[str, dict],
    product_lookup: Dict[str, dict],
    code_map: Dict[str, str],
    start_box_no: int,
) -> tuple[list[dict], int]:
    groups: List[dict] = []
    next_box_no = start_box_no

    fullbox_types = [
        ("single_fullboxes", ""),
        ("group_mixed_fullboxes", ""),
        ("tolerance_mixed_fullboxes", ""),
    ]

    for bucket_name, note in fullbox_types:
        for box in fullbox_result.get(bucket_name, []) or []:
            box_code = str(box.get("box_code", "") or "").strip()
            box_meta = box_lookup.get(box_code, {})
            outer_size_cm = box_meta.get("outer_size_cm", ())
            box_size_cm = _format_box_size(outer_size_cm)
            box_weight_kg = _to_float(box_meta.get("box_weight_kg"), 0.0)
            weight_mode = str(box_meta.get("weight_mode", "") or "").strip()
            items = box.get("items", []) or []

            if len(items) == 1:
                item = items[0]
                product_name = str(item.get("product_name", "") or "").strip()
                item_qty = _to_int(item.get("qty"), 0)
                unit_weight_kg = _to_float(product_lookup.get(product_name, {}).get("unit_weight_kg"), 0.0)
                if weight_mode == "fullbox_gross":
                    gross_weight = box_weight_kg
                else:
                    gross_weight = box_weight_kg + (item_qty * unit_weight_kg)

                groups.append(
                    _new_export_group(
                        box_size_cm=box_size_cm,
                        box_no=next_box_no,
                        box_count=1,
                        each_qty=item_qty,
                        weight_kg=gross_weight,
                        total_weight_kg=gross_weight,
                        item_rows=[
                            _new_item_row(
                                product_code=code_map.get(product_name, ""),
                                product_name=product_name,
                                packing_unit=item_qty,
                            )
                        ],
                        note=note,
                    )
                )
            else:
                total_each_qty = 0
                total_item_weight = 0.0
                item_rows = []

                for item in items:
                    product_name = str(item.get("product_name", "") or "").strip()
                    item_qty = _to_int(item.get("qty"), 0)
                    unit_weight_kg = _to_float(product_lookup.get(product_name, {}).get("unit_weight_kg"), 0.0)
                    total_each_qty += item_qty
                    total_item_weight += item_qty * unit_weight_kg
                    item_rows.append(
                        _new_item_row(
                            product_code=code_map.get(product_name, ""),
                            product_name=product_name,
                            packing_unit=item_qty,
                        )
                    )

                if weight_mode == "fullbox_gross":
                    gross_weight = box_weight_kg
                else:
                    gross_weight = box_weight_kg + total_item_weight
                groups.append(
                    _new_export_group(
                        box_size_cm=box_size_cm,
                        box_no=next_box_no,
                        box_count=1,
                        each_qty=total_each_qty,
                        weight_kg=gross_weight,
                        total_weight_kg=gross_weight,
                        item_rows=item_rows,
                        note=note,
                    )
                )

            next_box_no += 1

    return groups, next_box_no


def _build_repack_groups(
    final_result: Dict[str, Any],
    code_map: Dict[str, str],
    start_box_no: int,
) -> tuple[list[dict], int]:
    groups: List[dict] = []
    next_box_no = start_box_no

    for plan in final_result.get("final_plans", []) or []:
        product_name = str(plan.get("product_name", "") or "").strip()
        package_pack_qty = _to_int(plan.get("package_pack_qty"), 1)
        calc_unit_type = _normalize_calc_unit(plan.get("calc_unit_type"))
        outer_size_cm = plan.get("outer_size_cm", ())

        for box_line in plan.get("box_lines", []) or []:
            box_qty = _to_int(box_line.get("qty"), 0)
            if box_qty <= 0:
                continue

            if calc_unit_type == "package":
                each_qty = box_qty * package_pack_qty
                packing_unit = package_pack_qty
            else:
                each_qty = box_qty
                packing_unit = box_qty

            box_size_cm = _format_box_size(
                outer_size_cm,
                box_line.get("trimmed_outer_height_cm"),
            )

            groups.append(
                _new_export_group(
                    box_size_cm=box_size_cm,
                    box_no=next_box_no,
                    box_count=1,
                    each_qty=each_qty,
                    weight_kg=box_line.get("gross_weight_est"),
                    total_weight_kg=box_line.get("gross_weight_est"),
                    item_rows=[
                        _new_item_row(
                            product_code=code_map.get(product_name, ""),
                            product_name=product_name,
                            packing_unit=packing_unit,
                        )
                    ],
                    note="",
                )
            )
            next_box_no += 1

    return groups, next_box_no


def _build_fixed_box_mix_groups(
    fixed_box_result: Dict[str, Any],
    code_map: Dict[str, str],
) -> List[dict]:
    groups: List[dict] = []
    selected_box = fixed_box_result.get("selected_box", {}) or {}
    outer_size_cm = (
        selected_box.get("외경가로(cm)"),
        selected_box.get("외경세로(cm)"),
        selected_box.get("외경높이(cm)"),
    )
    box_size_cm = _format_box_size(outer_size_cm)

    for box in fixed_box_result.get("boxes", []) or []:
        item_rows = []
        each_qty = 0

        for item in box.get("items", []) or []:
            calc_unit_type = _normalize_calc_unit(item.get("calc_unit_type"))
            package_pack_qty = _to_int(item.get("package_pack_qty"), 1)
            alloc_qty = _to_int(item.get("qty"), 0)

            if calc_unit_type == "package":
                packing_unit = package_pack_qty
                line_each_qty = alloc_qty * package_pack_qty
            else:
                packing_unit = alloc_qty
                line_each_qty = alloc_qty

            each_qty += line_each_qty
            product_name = str(item.get("product_name", "") or "").strip()
            item_rows.append(
                _new_item_row(
                    product_code=code_map.get(product_name, ""),
                    product_name=product_name,
                    packing_unit=packing_unit,
                )
            )

        groups.append(
            _new_export_group(
                box_no=_to_int(box.get("box_no"), 0),
                box_size_cm=box_size_cm,
                box_count=1,
                each_qty=each_qty,
                weight_kg=box.get("gross_weight_est"),
                total_weight_kg=box.get("gross_weight_est"),
                item_rows=item_rows,
                note="",
            )
        )

    return groups


def _merge_consecutive_groups(groups: List[dict]) -> List[dict]:
    merged: List[dict] = []

    for group in groups:
        if not merged:
            merged.append(deepcopy(group))
            continue

        prev = merged[-1]
        same_key = (
            prev["box_size_cm"],
            prev["weight_kg"],
            prev["note"],
            prev["item_rows"],
        ) == (
            group["box_size_cm"],
            group["weight_kg"],
            group["note"],
            group["item_rows"],
        )

        if same_key and len(prev["item_rows"]) == 1 and prev["box_no_end"] + 1 == group["box_no_start"]:
            prev["box_no_end"] = group["box_no_end"]
            prev["box_count"] += group["box_count"]
            prev["each_qty"] += group["each_qty"]
            prev["total_weight_kg"] = _fmt_weight(
                _to_float(prev["total_weight_kg"], 0.0) + _to_float(group["total_weight_kg"], 0.0)
            )
            continue

        merged.append(deepcopy(group))

    return merged


def _box_no_display(group: Dict[str, Any]) -> str:
    start = _to_int(group.get("box_no_start"), 0)
    end = _to_int(group.get("box_no_end"), 0)

    if start <= 0:
        return ""
    if start == end:
        return str(start)
    return f"{start}~{end}"


def _build_reference_rows(groups: List[dict]) -> List[tuple]:
    seen = set()
    reference_rows = []

    for group in groups:
        for item in group.get("item_rows", []):
            key = (
                item["product_code"],
                item["product_name"],
                item["packing_unit"],
                group["box_size_cm"],
                group["weight_kg"],
            )
            if key in seen:
                continue

            seen.add(key)
            reference_rows.append(
                (
                    item["product_code"],
                    item["product_name"],
                    item["packing_unit"],
                    group["box_size_cm"],
                    group["weight_kg"],
                )
            )

    return reference_rows


def _resolve_shipment_mark(groups: List[dict], explicit_mark: str = "") -> str:
    mark = str(explicit_mark or "").strip()
    if mark:
        return mark

    product_codes = []
    for group in groups:
        for item in group.get("item_rows", []):
            code = str(item.get("product_code", "") or "").strip()
            if code:
                product_codes.append(code)

    unique_codes = sorted(set(product_codes))
    if len(unique_codes) == 1:
        return unique_codes[0]
    return ""


def _collect_issues(engine_result: Dict[str, Any]) -> List[tuple]:
    issues: List[tuple] = []

    sources = [
        ("fullbox_result", "not_found"),
        ("final_result", "no_box_fit"),
        ("final_result", "unresolved"),
        ("final_result", "invalid_specs"),
        ("repack_box_result", "no_box_fit"),
        ("repack_box_result", "unresolved"),
        ("repack_box_result", "invalid_specs"),
    ]

    for container_key, issue_key in sources:
        container = engine_result.get(container_key, {}) or {}
        for item in container.get(issue_key, []) or []:
            if not isinstance(item, dict):
                issues.append((container_key, issue_key, str(item), "", ""))
                continue

            issues.append(
                (
                    container_key,
                    issue_key,
                    str(item.get("product_name", "") or ""),
                    str(item.get("qty", "") or ""),
                    str(item.get("reason", "") or ""),
                )
            )

    for item in engine_result.get("match_result", []) or []:
        if not isinstance(item, dict):
            continue

        status = str(item.get("status", "") or "").strip().lower()
        if status not in {"ambiguous", "unresolved"}:
            continue

        issues.append(
            (
                "match_result",
                status,
                str(item.get("input_name", "") or item.get("raw_input", "") or ""),
                str(item.get("qty", "") or ""),
                str(item.get("reason", "") or item.get("message", "") or ""),
            )
        )

    unique = []
    seen = set()
    for row in issues:
        if row in seen:
            continue
        seen.add(row)
        unique.append(row)

    return unique


def _apply_sheet1_layout(ws, title: str) -> None:
    ws.title = "Sheet1"
    thin = Side(style="thin", color="000000")

    ws.merge_cells("B2:K2")
    ws["B2"] = title
    ws["B2"].font = Font(size=12, bold=False)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[2].height = 28

    header_fills = {
        2: "FFFFFF",
        3: "FFFFFF",
        4: "FFF2CC",
        5: "FFFFFF",
        6: "E2EFDA",
        7: "E2EFDA",
        8: "E2EFDA",
        9: "FFFFFF",
        10: "FFFFFF",
        11: "FFF2CC",
    }

    for idx, header in enumerate(HEADER_TITLES, start=2):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=header_fills.get(idx, "FFFFFF"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[3].height = 30

    widths = {
        "A": 4,
        "B": 14,
        "C": 12,
        "D": 30,
        "E": 12,
        "F": 20,
        "G": 10,
        "H": 12,
        "I": 10,
        "J": 12,
        "K": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "B4"


def _style_data_cell(cell, fill_color: str) -> None:
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_sheet1_rows(ws, groups: List[dict], shipment_mark: str = "") -> None:
    band_colors = ["F2F2F2", "FCE4D6", "D9E2F3", "E2F0D9"]
    row_idx = 4
    data_start_row = row_idx
    total_box_count = 0
    total_each_qty = 0
    total_weight = 0.0

    for group_idx, group in enumerate(groups):
        fill_color = band_colors[group_idx % len(band_colors)]
        item_rows = group.get("item_rows", []) or [{}]
        span = max(1, len(item_rows))
        start_row = row_idx
        end_row = row_idx + span - 1

        for current_row in range(start_row, end_row + 1):
            ws.row_dimensions[current_row].height = 26

        shared_values = {
            3: _box_no_display(group),
            6: group["box_size_cm"],
            7: group["box_count"],
            8: group["each_qty"],
            9: group["weight_kg"],
            10: group["total_weight_kg"],
            11: group["note"],
        }

        for offset, item in enumerate(item_rows):
            current_row = start_row + offset
            item_values = {
                4: item.get("product_name", ""),
                5: item.get("packing_unit", ""),
            }

            for col_idx, value in item_values.items():
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                _style_data_cell(cell, fill_color)
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, value in shared_values.items():
            cell = ws.cell(row=start_row, column=col_idx, value=value)
            _style_data_cell(cell, fill_color)
            if col_idx in {9, 10} and value not in ("", None):
                cell.number_format = "0.###"
            if col_idx == 11:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if span > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=end_row,
                    end_column=col_idx,
                )

        total_box_count += _to_int(group.get("box_count"), 0)
        total_each_qty += _to_int(group.get("each_qty"), 0)
        total_weight += _to_float(group.get("total_weight_kg"), 0.0)
        row_idx = end_row + 1

    data_end_row = row_idx - 1

    if data_end_row > data_start_row:
        ws.merge_cells(start_row=data_start_row, start_column=2, end_row=data_end_row, end_column=2)
    if data_end_row >= data_start_row:
        shipment_cell = ws.cell(row=data_start_row, column=2, value=shipment_mark)
        _style_data_cell(shipment_cell, "FFFFFF")
        shipment_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row = row_idx
    total_fill = "FFFFFF"

    for col_idx in range(2, 12):
        cell = ws.cell(row=total_row, column=col_idx, value="")
        _style_data_cell(cell, total_fill)

    ws.cell(row=total_row, column=3, value="총합계")
    ws.cell(row=total_row, column=7, value=total_box_count)
    ws.cell(row=total_row, column=8, value=total_each_qty)
    ws.cell(row=total_row, column=10, value=round(total_weight, 3))
    ws.cell(row=total_row, column=10).number_format = "0.###"
    for col_idx in (3, 7, 8, 10):
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
    ws.row_dimensions[total_row].height = 26


def _write_reference_sheet(wb: Workbook, reference_rows: List[tuple]) -> None:
    ws = wb.create_sheet("Sheet2")
    headers = ("상품코드", "상품명", "입수량", "사이즈", "무게")
    ws.append(headers)

    for row in reference_rows:
        ws.append(row)

    for column, width in {"A": 14, "B": 30, "C": 10, "D": 20, "E": 10}.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"


def _write_issues_sheet(wb: Workbook, issues: List[tuple]) -> None:
    if not issues:
        return

    ws = wb.create_sheet("Issues")
    ws.append(("source", "issue_type", "product_name", "qty", "reason"))

    for row in issues:
        ws.append(row)

    for column, width in {"A": 16, "B": 16, "C": 28, "D": 10, "E": 28}.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"


def export_engine_result_to_sulu_med_xlsx(
    engine_result: Dict[str, Any],
    output_path: str | Path,
    title: str = "Sulu Med 출하건",
    master_path: str | Path | None = None,
    shipment_mark: str = "",
) -> Dict[str, Any]:
    context = _build_master_context(master_path=master_path)
    product_lookup = context["product_lookup"]
    box_lookup = context["box_lookup"]
    code_map = _build_code_map(engine_result, product_lookup)

    groups: List[dict] = []
    next_box_no = 1

    fullbox_groups, next_box_no = _build_fullbox_groups(
        fullbox_result=engine_result.get("fullbox_result", {}) or {},
        box_lookup=box_lookup,
        product_lookup=product_lookup,
        code_map=code_map,
        start_box_no=next_box_no,
    )
    groups.extend(fullbox_groups)

    repack_groups, next_box_no = _build_repack_groups(
        final_result=engine_result.get("final_result", {}) or {},
        code_map=code_map,
        start_box_no=next_box_no,
    )
    groups.extend(repack_groups)

    merged_groups = _merge_consecutive_groups(groups)
    resolved_mark = _resolve_shipment_mark(merged_groups, explicit_mark=shipment_mark)
    reference_rows = _build_reference_rows(merged_groups)
    issues = _collect_issues(engine_result)

    wb = Workbook()
    ws = wb.active
    _apply_sheet1_layout(ws, title=title)
    _write_sheet1_rows(ws, merged_groups, shipment_mark=resolved_mark)
    _write_reference_sheet(wb, reference_rows)
    _write_issues_sheet(wb, issues)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    return {
        "output_path": str(output_file),
        "row_count": len(merged_groups),
        "issue_count": len(issues),
    }


def build_router_display_payload(
    router_result: Dict[str, Any],
    master_path: str | Path | None = None,
    shipment_mark: str = "",
) -> Dict[str, Any]:
    selected_engine = str(router_result.get("selected_engine", "") or "").strip()
    result = router_result.get("result", {}) or {}

    if selected_engine == "run_packing_engine":
        context = _build_master_context(master_path=master_path)
        product_lookup = context["product_lookup"]
        box_lookup = context["box_lookup"]
        code_map = _build_code_map(result, product_lookup)

        groups: List[dict] = []
        next_box_no = 1

        fullbox_groups, next_box_no = _build_fullbox_groups(
            fullbox_result=result.get("fullbox_result", {}) or {},
            box_lookup=box_lookup,
            product_lookup=product_lookup,
            code_map=code_map,
            start_box_no=next_box_no,
        )
        groups.extend(fullbox_groups)

        repack_groups, next_box_no = _build_repack_groups(
            final_result=result.get("final_result", {}) or {},
            code_map=code_map,
            start_box_no=next_box_no,
        )
        groups.extend(repack_groups)

        merged_groups = _merge_consecutive_groups(groups)
        resolved_mark = _resolve_shipment_mark(merged_groups, explicit_mark=shipment_mark)
        reference_rows = _build_reference_rows(merged_groups)
        issues = _collect_issues(result)

        return {
            "selected_engine": selected_engine,
            "groups": merged_groups,
            "shipment_mark": resolved_mark,
            "reference_rows": reference_rows,
            "issues": issues,
        }

    context = _build_master_context(master_path=master_path)
    product_lookup = context["product_lookup"]
    code_map = {
        product_name: payload.get("product_code", "")
        for product_name, payload in product_lookup.items()
        if payload.get("product_code")
    }

    groups: List[dict] = []
    issues: List[tuple] = []

    if selected_engine == "fixed_box_mix_checker":
        groups = _build_fixed_box_mix_groups(result, code_map=code_map)

        for issue_key in ("not_found", "unresolved", "invalid_specs", "no_fit"):
            for item in result.get(issue_key, []) or []:
                if isinstance(item, dict):
                    issues.append(
                        (
                            selected_engine,
                            issue_key,
                            str(item.get("product_name", "") or ""),
                            str(item.get("qty", "") or ""),
                            str(item.get("reason", "") or ""),
                        )
                    )

        if bool(result.get("pack_failed", False)):
            issues.append(
                (
                    selected_engine,
                    "pack_failed",
                    "",
                    "",
                    str(result.get("pack_failed_reason", "") or ""),
                )
            )
    else:
        raise ValueError(f"지원하지 않는 라우터 결과입니다: {selected_engine or 'UNKNOWN'}")

    merged_groups = _merge_consecutive_groups(groups)
    resolved_mark = _resolve_shipment_mark(merged_groups, explicit_mark=shipment_mark)
    reference_rows = _build_reference_rows(merged_groups)

    return {
        "selected_engine": selected_engine,
        "groups": merged_groups,
        "shipment_mark": resolved_mark,
        "reference_rows": reference_rows,
        "issues": issues,
    }


def export_router_result_to_sulu_med_xlsx(
    router_result: Dict[str, Any],
    output_path: str | Path,
    title: str = "Sulu Med 출하건",
    master_path: str | Path | None = None,
    shipment_mark: str = "",
) -> Dict[str, Any]:
    payload = build_router_display_payload(
        router_result=router_result,
        master_path=master_path,
        shipment_mark=shipment_mark,
    )

    groups = payload["groups"]
    resolved_mark = payload["shipment_mark"]
    reference_rows = payload["reference_rows"]
    issues = payload["issues"]

    wb = Workbook()
    ws = wb.active
    _apply_sheet1_layout(ws, title=title)
    _write_sheet1_rows(ws, groups, shipment_mark=resolved_mark)
    _write_reference_sheet(wb, reference_rows)
    _write_issues_sheet(wb, issues)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    return {
        "output_path": str(output_file),
        "row_count": len(groups),
        "issue_count": len(issues),
        "selected_engine": payload["selected_engine"],
    }
